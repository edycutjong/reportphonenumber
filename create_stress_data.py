import openpyxl
import random
from datetime import datetime, timedelta

def generate_stress_data(num_records=500):
    providers = ["Telkomsel", "Indosat Ooredoo", "XL Axiata", "Smartfren", "Tri (3)"]
    statuses = ["Sukses", "Pending", "Gagal"]
    status_weights = [0.85, 0.10, 0.05] # mostly success, some pending, some failure

    # Generate records
    records = []
    base_time = datetime(2026, 6, 26, 9, 0, 0) # starting at 9:00 AM

    for i in range(1, num_records + 1):
        # Generate phone number (avoid scientific notation by formatting as text)
        prefix = random.choice(["0811", "0812", "0813", "0821", "0852", "0853", "0817", "0818", "0819", "0896", "0897"])
        suffix = "".join([str(random.randint(0, 9)) for _ in range(8)])
        phone_num = f"{prefix}{suffix}"

        # Increment time randomly by 1 to 30 seconds
        base_time += timedelta(seconds=random.randint(1, 30))
        date_str = base_time.strftime("%d/%m/%Y")
        time_str = base_time.strftime("%H:%M:%S")

        provider = random.choice(providers)
        serial = f"SN{2026000000 + i}"
        status = random.choices(statuses, weights=status_weights)[0]

        records.append([
            i,          # No
            phone_num,  # No Pelanggan
            date_str,   # Tanggal
            time_str,   # Waktu
            provider,   # Provider
            serial,     # Serial Number
            status      # Status
        ])

    return records

def save_to_xlsx(records, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stress Test 500"

    headers = ["No", "No Pelanggan", "Tanggal", "Waktu", "Provider", "Serial Number", "Status"]
    ws.append(headers)

    for row in records:
        ws.append(row)

    # Force format all cells in "No Pelanggan" and "Serial Number" as text to prevent excel formatting issues
    for row in range(2, len(records) + 2):
        ws.cell(row=row, column=2).number_format = '@' # No Pelanggan
        ws.cell(row=row, column=6).number_format = '@' # Serial Number

    wb.save(filename)
    print(f"Excel stress test file '{filename}' generated with {len(records)} records.")

def save_to_csv(records, filename):
    import csv
    headers = ["No", "No Pelanggan", "Tanggal", "Waktu", "Provider", "Serial Number", "Status"]
    with open(filename, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(records)
    print(f"CSV stress test file '{filename}' generated with {len(records)} records.")

if __name__ == "__main__":
    records = generate_stress_data(500)
    save_to_xlsx(records, "/Users/edycu/Playground/ReportPhoneNumber/test_stress.xlsx")
    save_to_csv(records, "/Users/edycu/Playground/ReportPhoneNumber/test_stress.csv")
